from Demos.SystemParametersInfo import new_value
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import time

def update_excel_data(file_path, search_term, col_name, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}
    # update the cell
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == col_name:
            Dict['col'] = i
    for i in range(1,sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == search_term:
                Dict['row'] = i


    sheet.cell(row=Dict['row'], column=Dict['col']).value = new_value
    book.save(file_path)


file_path = "C:/Users/SinhaShreya/Downloads/download.xlsx"
fruit = "Apple"
newValue = "957"
driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

#click on download button
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()

# edit the excel with updated value
update_excel_data(file_path, fruit, "price", newValue)

# click on choose file to upload updated excel sheet
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")

file_input.send_keys(file_path)
wait = WebDriverWait(driver,10)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
time.sleep(5)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute('data-column-id')
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text

assert actual_price == new_value
#close browser
driver.quit()